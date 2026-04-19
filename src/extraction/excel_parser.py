"""
Excel MASTER sheet parser for corporate credit rating files.

Handles the non-standard key-value structure of MASTER sheets where:
- Column 0 is always None (spacer)
- Column 1 contains field labels
- Column 2+ contains values (some fields span multiple columns for multi-industry data)

The parser extracts all company metadata, risk profiles, and credit metrics
into a standardized dictionary format.
"""

from __future__ import annotations

import hashlib
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl

logger = logging.getLogger(__name__)


# ============================================================================
# FIELD MAPPING: Maps row labels to standardized field names
# ============================================================================

# Key-value fields: label → field_name (value in column 2)
SINGLE_VALUE_FIELDS = {
    "Rated entity": "company_name",
    "CorporateSector": "sector",
    "Segmentation criteria": "segmentation_criteria",
    "Reporting Currency/Units": "currency",
    "Country of origin": "country",
    "Accounting principles": "accounting_principles",
    "End of business year": "year_end_month",
    # Business risk profile
    "Business risk profile": "business_risk_profile",
    "(Blended) Industry risk profile": "blended_industry_risk_profile",
    "Competitive Positioning": "competitive_positioning",
    "Market share": "market_share",
    "Diversification": "diversification",
    "Operating profitability": "operating_profitability",
    "Sector/company-specific factors (1)": "sector_specific_factor_1",
    "Sector/company-specific factors (2)": "sector_specific_factor_2",
    # Financial risk profile
    "Financial risk profile": "financial_risk_profile",
    "Leverage": "leverage",
    "Interest cover": "interest_cover",
    "Cash flow cover": "cash_flow_cover",
}

# Multi-value fields: these span across columns 2+ for multiple industries
MULTI_VALUE_LABELS = {
    "Rating methodologies applied",
    "Industry risk",
    "Industry risk score",
    "Industry weight",
}

# Credit metrics section
CREDIT_METRICS_HEADER = "[Scope Credit Metrics]"
CREDIT_METRIC_NAMES = {
    "Scope-adjusted EBITDA interest cover",
    "Scope-adjusted debt/EBITDA",
    "Scope-adjusted FFO/debt",
    "Scope-adjusted loan/value",
    "Scope-adjusted FOCF/debt",
    "Liquidity",
}


def compute_file_hash(file_path: str | Path) -> str:
    """Compute SHA-256 hash of a file for deduplication."""
    sha256 = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha256.update(chunk)
    return sha256.hexdigest()


def extract_version_from_filename(filename: str) -> tuple[str, int]:
    """
    Extract company identifier and version number from filename.
    
    E.g., 'corporates_A_1.xlsm' → ('company_A', 1)
         'corporates_B_2.xlsm' → ('company_B', 2)
    """
    stem = Path(filename).stem  # e.g., 'corporates_A_1'
    parts = stem.split("_")
    
    if len(parts) >= 3:
        # company_id from the letter/identifier parts
        company_letter = parts[-2]  # 'A' or 'B'
        version = int(parts[-1])    # 1 or 2
        company_id = f"company_{company_letter}"
        return company_id, version
    
    # Fallback: use full stem as company_id
    return stem, 1


def parse_master_sheet(file_path: str | Path) -> dict[str, Any]:
    """
    Parse the MASTER sheet from an Excel .xlsm file.
    
    The MASTER sheet has a non-standard key-value structure:
    - Row index 0: column headers (all "Unnamed: X") — skipped
    - Subsequent rows: [None, label, value, optional_value_2, ...]
    
    Returns a standardized dictionary with all extracted fields,
    file metadata, and data lineage information.
    
    Args:
        file_path: Path to the .xlsm Excel file
        
    Returns:
        Dictionary containing all extracted company data
        
    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If MASTER sheet is missing or unparseable
    """
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    
    logger.info("Parsing MASTER sheet from: %s", file_path.name)
    
    # Load workbook (read_only for performance, data_only for computed values)
    wb = openpyxl.load_workbook(
        str(file_path),
        read_only=True,
        keep_vba=False,
        data_only=True,
    )
    
    if "MASTER" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"MASTER sheet not found in {file_path.name}. Available: {wb.sheetnames}")
    
    ws = wb["MASTER"]
    
    # Read all rows into memory (small sheet — ~40 meaningful rows)
    rows: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(tuple(row))
    wb.close()
    
    if not rows:
        raise ValueError(f"MASTER sheet is empty in {file_path.name}")
    
    # ---- Extract fields ----
    result: dict[str, Any] = {}
    rating_methodologies: list[str] = []
    industry_risks: list[dict[str, Any]] = []
    industry_risk_industries: list[str] = []
    industry_risk_scores: list[str] = []
    industry_risk_weights: list[float] = []
    credit_metrics: dict[str, dict[str, Any]] = {}
    credit_metric_years: list[str] = []
    liquidity_rating: Optional[str] = None  # Distinguish from credit metric Liquidity
    
    for row_idx, row in enumerate(rows):
        if not row or len(row) < 2:
            continue
        
        # Column 1 is the label (index 1 in 0-based)
        label = row[1]
        if label is None:
            continue
        
        label = str(label).strip()
        
        # ---- Single-value fields ----
        if label in SINGLE_VALUE_FIELDS:
            field_name = SINGLE_VALUE_FIELDS[label]
            value = row[2] if len(row) > 2 else None
            if value is not None:
                result[field_name] = str(value).strip() if isinstance(value, str) else value
            else:
                result[field_name] = None
                
        # ---- Liquidity (rating context, row ~30) ----
        elif label == "Liquidity" and credit_metric_years == []:
            # This is the rating-context liquidity (e.g., "+1 notch"), not the credit metric
            value = row[2] if len(row) > 2 else None
            result["liquidity"] = str(value).strip() if value is not None else None
        
        # ---- Rating methodologies (multi-value across columns) ----
        elif label == "Rating methodologies applied":
            for col_idx in range(2, len(row)):
                val = row[col_idx]
                if val is not None and str(val).strip():
                    rating_methodologies.append(str(val).strip())
        
        # ---- Industry risk names ----
        elif label == "Industry risk":
            for col_idx in range(2, len(row)):
                val = row[col_idx]
                if val is not None and str(val).strip():
                    industry_risk_industries.append(str(val).strip())
        
        # ---- Industry risk scores ----
        elif label == "Industry risk score":
            for col_idx in range(2, len(row)):
                val = row[col_idx]
                if val is not None and str(val).strip():
                    industry_risk_scores.append(str(val).strip())
        
        # ---- Industry risk weights ----
        elif label == "Industry weight":
            for col_idx in range(2, len(row)):
                val = row[col_idx]
                if val is not None:
                    try:
                        industry_risk_weights.append(float(val))
                    except (ValueError, TypeError):
                        pass
        
        # ---- Credit metrics header (year columns) ----
        elif label == CREDIT_METRICS_HEADER:
            for col_idx in range(2, len(row)):
                val = row[col_idx]
                if val is not None:
                    credit_metric_years.append(str(val).strip())
        
        # ---- Credit metric data rows ----
        elif label in CREDIT_METRIC_NAMES or (
            credit_metric_years and label == "Liquidity"
        ):
            metric_data: dict[str, Any] = {}
            for yr_idx, year in enumerate(credit_metric_years):
                col_idx = yr_idx + 2
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is not None and str(val).strip().lower() not in ("locked",):
                        if str(val).strip().lower() == "no data":
                            metric_data[year] = None
                        else:
                            try:
                                metric_data[year] = float(val)
                            except (ValueError, TypeError):
                                metric_data[year] = str(val).strip()
                    # Skip "Locked" values
            
            metric_name = label
            if label == "Liquidity" and credit_metric_years:
                metric_name = "Liquidity (metric)"
            credit_metrics[metric_name] = metric_data
    
    # ---- Assemble industry risk objects ----
    max_industries = max(
        len(industry_risk_industries),
        len(industry_risk_scores),
        len(industry_risk_weights),
    ) if industry_risk_industries else 0
    
    for i in range(max_industries):
        risk_entry: dict[str, Any] = {}
        if i < len(industry_risk_industries):
            risk_entry["industry"] = industry_risk_industries[i]
        if i < len(industry_risk_scores):
            risk_entry["score"] = industry_risk_scores[i]
        if i < len(industry_risk_weights):
            risk_entry["weight"] = industry_risk_weights[i]
        industry_risks.append(risk_entry)
    
    # ---- File metadata & lineage ----
    file_hash = compute_file_hash(file_path)
    company_id, version = extract_version_from_filename(file_path.name)
    
    result["rating_methodologies"] = rating_methodologies
    result["industry_risks"] = industry_risks
    result["credit_metrics"] = credit_metrics
    
    # File-level metadata
    result["_metadata"] = {
        "source_file": file_path.name,
        "file_path": str(file_path.absolute()),
        "file_hash": file_hash,
        "file_size_bytes": file_path.stat().st_size,
        "extracted_at": datetime.utcnow().isoformat(),
        "company_id": company_id,
        "version": version,
        "sheet_name": "MASTER",
        "total_rows_parsed": len(rows),
    }
    
    logger.info(
        "Extracted data from %s: company=%s, sector=%s, version=%d",
        file_path.name,
        result.get("company_name", "UNKNOWN"),
        result.get("sector", "UNKNOWN"),
        version,
    )
    
    return result


def extract_all_files(data_dir: str | Path) -> list[dict[str, Any]]:
    """
    Extract MASTER sheet data from all .xlsm files in a directory.
    
    Args:
        data_dir: Directory containing .xlsm files
        
    Returns:
        List of extracted data dictionaries
    """
    data_dir = Path(data_dir)
    
    if not data_dir.exists():
        raise FileNotFoundError(f"Data directory not found: {data_dir}")
    
    xlsm_files = sorted(data_dir.glob("*.xlsm"))
    
    if not xlsm_files:
        logger.warning("No .xlsm files found in %s", data_dir)
        return []
    
    logger.info("Found %d .xlsm files in %s", len(xlsm_files), data_dir)
    
    results: list[dict[str, Any]] = []
    for file_path in xlsm_files:
        try:
            data = parse_master_sheet(file_path)
            results.append(data)
        except Exception as e:
            logger.error("Failed to extract %s: %s", file_path.name, e)
            results.append({
                "_metadata": {
                    "source_file": file_path.name,
                    "file_path": str(file_path.absolute()),
                    "file_hash": compute_file_hash(file_path),
                    "file_size_bytes": file_path.stat().st_size,
                    "extracted_at": datetime.utcnow().isoformat(),
                    "error": str(e),
                },
                "_extraction_error": str(e),
            })
    
    return results
